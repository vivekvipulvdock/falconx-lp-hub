#!/usr/bin/env python3
"""
Regenerates FalconX_LP_Hub.html from FalconX_LP_Hub_Database.xlsx.

Usage:
    python3 rebuild_site.py [source_html] [database_xlsx] [output_html]

Defaults:
    source_html   = FalconX_LP_Hub.html
    database_xlsx = FalconX_LP_Hub_Database.xlsx
    output_html   = FalconX_LP_Hub.html   (overwritten in place)

Only the data between the "/* ===DATABASE_START=== */" and
"/* ===DATABASE_END=== */" markers inside the page's script is replaced.
Everything else — design, layout, CSS, fonts, embedded assets, and all
interactive logic — is read verbatim from source_html and left untouched.
"""
import sys
import json
from openpyxl import load_workbook

CORE_FIELDS = ['id', 'name', 'mono', 'hue', 'sector', 'stage', 'status', 'inv',
               'val', 'own', 'date', 'round', 'post', 'upd', 'desc', 'thesis']
SHEET_CORE_FIELDS = ['id', 'name', 'mono', 'sector', 'stage', 'status', 'inv',
                     'val', 'own', 'date', 'round', 'post', 'upd', 'desc',
                     'thesis', 'summary']
FUND_KEYS = ['name', 'size', 'deployed', 'companies', 'value', 'invested',
             'exits', 'tvpi', 'dpi', 'irr', 'positioning']

START_MARKER = '/* ===DATABASE_START=== */'
END_MARKER = '/* ===DATABASE_END=== */'


def js_str(s):
    s = str(s)
    s = s.replace('\\', '\\\\').replace("'", "\\'")
    return "'" + s + "'"


def js_val(v):
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, str):
        return js_str(v)
    return repr(v)


def rows(ws, min_row=2):
    out = []
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        if row is None or row[0] is None:
            continue
        out.append(row)
    return out


def read_database(xlsx_path, original_hue_by_id):
    wb = load_workbook(xlsx_path, data_only=True)

    companies = {}
    order = []
    next_hue_pool = [h for h in range(0, 360, 24)]
    pool_i = 0
    for r in rows(wb['Companies']):
        rec = dict(zip(SHEET_CORE_FIELDS, r))
        cid = str(rec['id']).strip()
        order.append(cid)
        hue = original_hue_by_id.get(cid)
        if hue is None:
            hue = next_hue_pool[pool_i % len(next_hue_pool)]
            pool_i += 1
        companies[cid] = {
            'id': cid, 'name': rec['name'], 'mono': rec['mono'], 'hue': hue,
            'sector': rec['sector'], 'stage': rec['stage'], 'status': rec['status'],
            'inv': float(rec['inv']), 'val': float(rec['val']), 'own': float(rec['own']),
            'date': rec['date'], 'round': rec['round'], 'post': float(rec['post']),
            'upd': float(rec['upd']), 'desc': rec['desc'], 'thesis': rec['thesis'],
            'summary': rec['summary'],
            'kpis': [], 'events': [], 'news': [], 'highlights': [],
            'internalNotes': [], 'docs': [],
        }

    for r in rows(wb['Company KPIs']):
        cid, _name, _order, label, value, delta, direction, conf, source = r
        companies[str(cid)]['kpis'].append({
            'label': label, 'value': str(value), 'delta': delta,
            'dir': direction, 'conf': conf, 'source': source,
        })

    for r in rows(wb['Company Timeline']):
        cid, _name, _order, date, typ, title = r
        companies[str(cid)]['events'].append([date, typ, title])

    for r in rows(wb['Company News']):
        cid, _name, _order, date, src, title = r
        companies[str(cid)]['news'].append([date, src, title])

    for r in rows(wb['Company Highlights']):
        cid, _name, _order, text = r
        companies[str(cid)]['highlights'].append(text)

    for r in rows(wb['Company Internal Notes']):
        cid, _name, _order, tone, text = r
        companies[str(cid)]['internalNotes'].append({'tone': tone, 'text': text})

    for r in rows(wb['Company Documents']):
        cid, _name, _order, label, via, internal = r
        is_internal = str(internal).strip().lower() in ('yes', 'true', '1')
        companies[str(cid)]['docs'].append({'label': label, 'via': via, 'internal': is_internal})

    company_list = [companies[cid] for cid in order]

    fund_rows = rows(wb['Fund Overview'])
    fund = {}
    for k, (_label, v) in zip(FUND_KEYS, fund_rows):
        fund[k] = v
    fund['size'] = int(fund['size'])
    fund['deployed'] = float(fund['deployed'])
    fund['companies'] = int(fund['companies'])
    fund['value'] = float(fund['value'])
    fund['invested'] = float(fund['invested'])
    fund['exits'] = int(fund['exits'])
    fund['tvpi'] = float(fund['tvpi'])
    fund['dpi'] = float(fund['dpi'])

    ws9 = wb['Newsletter']
    meta_rows = list(ws9.iter_rows(min_row=2, max_row=6, max_col=2, values_only=True))
    meta = {k: v for k, v in zip(['issue', 'period', 'date', 'intro', 'signoff'], [r[1] for r in meta_rows])}

    numbers, notable = [], []
    mode = None
    for r in ws9.iter_rows(min_row=7, max_col=2, values_only=True):
        a, b = r[0], r[1]
        if a in ('Numbers Strip', 'Notable Items'):
            mode = None
            continue
        if a == 'Label' and b == 'Value':
            mode = 'numbers'
            continue
        if a == 'Tag' and b == 'Text':
            mode = 'notable'
            continue
        if a is None and b is None:
            continue
        if mode == 'numbers':
            numbers.append({'k': a, 'v': b})
        elif mode == 'notable':
            notable.append({'tag': a, 'text': b})

    highlight_ids = []
    for r in rows(wb['Newsletter Highlights']):
        cid, _name, _order, line, metric, mult = r
        highlight_ids.append([str(cid), line, metric, mult])

    return company_list, fund, meta, numbers, notable, highlight_ids


def render_core(companies):
    lines = []
    for c in companies:
        vals = [js_val(c[f]) for f in CORE_FIELDS]
        lines.append('      [' + ','.join(vals) + ']')
    return 'const core=[\n' + ',\n'.join(lines) + '\n    ];'


def render_kpi(k):
    return "{label:%s,value:%s,delta:%s,dir:%s,conf:%s,source:%s}" % (
        js_str(k['label']), js_str(k['value']), js_str(k['delta']),
        js_str(k['dir']), js_str(k['conf']), js_str(k['source']))


def render_event(e):
    date, typ, title = e
    return "[%s,%s,%s]" % (js_str(date), js_str(typ), js_str(title))


def render_news(n):
    date, src, title = n
    return "[%s,%s,%s]" % (js_str(date), js_str(src), js_str(title))


def render_note(n):
    return "{tone:%s,text:%s}" % (js_str(n['tone']), js_str(n['text']))


def render_doc(d):
    return "{label:%s,via:%s,internal:%s}" % (js_str(d['label']), js_str(d['via']), js_val(d['internal']))


def render_extra(companies):
    out = []
    for c in companies:
        kpis = ',\n        '.join(render_kpi(k) for k in c['kpis'])
        events = ',\n        '.join(render_event(e) for e in c['events'])
        news = ',\n        '.join(render_news(n) for n in c['news'])
        highlights = ',\n        '.join(js_str(h) for h in c['highlights'])
        notes = ',\n        '.join(render_note(n) for n in c['internalNotes'])
        docs = ',\n        '.join(render_doc(d) for d in c['docs'])
        entry = (
            "      %s:{\n"
            "        kpis:[\n        %s\n        ],\n"
            "        events:[\n        %s\n        ],\n"
            "        news:[\n        %s\n        ],\n"
            "        highlights:[\n        %s\n        ],\n"
            "        internalNotes:[\n        %s\n        ],\n"
            "        docs:[\n        %s\n        ],\n"
            "        summary:%s\n"
            "      }"
        ) % (js_str(c['id']), kpis, events, news, highlights, notes, docs, js_str(c['summary']))
        out.append(entry)
    return 'const EXTRA={\n' + ',\n'.join(out) + '\n    };'


def render_fund(fund):
    parts = [f'{k}:{js_val(fund[k])}' for k in FUND_KEYS]
    line1 = ', '.join(parts[:4])
    line2 = ', '.join(parts[4:10])
    line3 = ', '.join(parts[10:])
    return 'const fund={\n      ' + line1 + ',\n      ' + line2 + ',\n      ' + line3 + '\n    };'


def render_newsletter_meta(meta):
    return ("const newsletterMeta={issue:%s,period:%s,date:%s,intro:%s,signoff:%s};" %
            (js_str(meta['issue']), js_str(meta['period']), js_str(meta['date']),
             js_str(meta['intro']), js_str(meta['signoff'])))


def render_newsletter_numbers(numbers):
    items = ',\n      '.join('{k:%s,v:%s}' % (js_str(n['k']), js_str(n['v'])) for n in numbers)
    return 'const newsletterNumbers=[\n      ' + items + '\n    ];'


def render_newsletter_notable(notable):
    items = ',\n      '.join('{tag:%s,text:%s}' % (js_str(n['tag']), js_str(n['text'])) for n in notable)
    return 'const newsletterNotable=[\n      ' + items + '\n    ];'


def render_newsletter_highlights(highlight_ids):
    items = ['[%s,%s,%s,%s]' % (js_str(cid), js_str(line), js_str(metric), js_str(mult))
             for cid, line, metric, mult in highlight_ids]
    return 'const newsletterHighlightIds=[\n      ' + ',\n      '.join(items) + '\n    ];'


def build_database_block(companies, fund, meta, numbers, notable, highlight_ids):
    return '\n    '.join([
        START_MARKER,
        render_core(companies),
        render_extra(companies),
        render_fund(fund),
        render_newsletter_meta(meta),
        render_newsletter_numbers(numbers),
        render_newsletter_notable(notable),
        render_newsletter_highlights(highlight_ids),
        END_MARKER,
    ])


def main():
    source_html = sys.argv[1] if len(sys.argv) > 1 else 'FalconX_LP_Hub.html'
    db_xlsx = sys.argv[2] if len(sys.argv) > 2 else 'FalconX_LP_Hub_Database.xlsx'
    output_html = sys.argv[3] if len(sys.argv) > 3 else source_html

    with open(source_html, 'r', encoding='utf-8') as f:
        bundle_lines = f.readlines()

    template_line_idx = None
    for i, line in enumerate(bundle_lines):
        if line.startswith('"<!DOCTYPE html>'):
            template_line_idx = i
            break
    if template_line_idx is None:
        raise RuntimeError('Could not find the template data line in ' + source_html)

    template_html = json.loads(bundle_lines[template_line_idx])

    s = template_html.find(START_MARKER)
    e = template_html.find(END_MARKER)
    if s == -1 or e == -1:
        raise RuntimeError('Could not locate database markers in template')
    e += len(END_MARKER)

    import re
    import ast
    old_core_m = re.search(r"const core=(\[.*?\]);", template_html[s:e], re.S)
    original_core = ast.literal_eval(old_core_m.group(1))
    original_hue_by_id = {c[0]: c[3] for c in original_core}

    companies, fund, meta, numbers, notable, highlight_ids = read_database(db_xlsx, original_hue_by_id)

    new_block = build_database_block(companies, fund, meta, numbers, notable, highlight_ids)
    new_template = template_html[:s] + new_block + template_html[e:]

    encoded = json.dumps(new_template, ensure_ascii=False)
    encoded = encoded.replace('</', '<\\u002F')
    bundle_lines[template_line_idx] = encoded + '\n'

    with open(output_html, 'w', encoding='utf-8') as f:
        f.writelines(bundle_lines)

    print(f'Wrote {output_html} ({len(companies)} companies)')


if __name__ == '__main__':
    main()
